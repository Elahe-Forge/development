# import collections
# import json
import re
from io import BytesIO

import boto3
import helpers.utils as utils
import pandas as pd
from helpers.fields import other_fields, precise_fields, raw_fields

# from helpers.readers import DocumentReader
from openpyxl import load_workbook
from openpyxl.comments import Comment

# Define a regex pattern to match preferred share names - used in precise_output_df and support_output_df
regex_pattern = r"\b(?:[A-Z]+(?:\d+)?(?:\([A-Z0-9]+\))?(?:-[A-Z0-9]+)*|Seed|(\d+))\b"


def generate_precise_df(bucket, json_data):
    # Extract model ID from the JSON data
    model_id = json_data["model_id"]

    path_key = "precise_output_path"  # Key for the precise output path in JSON
    columns = []  # List to store column names for the DataFrame
    output_dict = {}  # Dictionary to store output data
    preferred_shares_list = None  # Initialize preferred shares list as None

    # Iterate over each field and its corresponding dictionary in precise_fields
    for field, dict in precise_fields.items():
        columns.append(field)  # Add field name to columns list

        output_file = dict["output_file"]  # Get the output file name

        file_path = json_data[output_file][path_key]  # Construct file path
        file_content = utils.load_s3_file(
            bucket, file_path
        )  # Load file content from S3

        json_obj = utils.convert_file_to_json_obj(
            model_id, file_content
        )  # Convert file content to JSON object

        # Handle "shares_type" field differently
        if field == "shares_type":
            preferred_shares_list = json_obj[
                dict["key_name"]
            ]  # Set preferred shares list

            # Initialize output_dict with share types
            for share_type in json_obj[dict["key_name"]]:
                output_dict[share_type] = [share_type]

        else:
            # Process other fields
            for pref_share_name in preferred_shares_list:
                pattern = re.compile(regex_pattern)  # Compile regex pattern
                pattern_pref_share_name = pattern.search(
                    pref_share_name
                ).group()  # Extract the pattern match

                matched = 0
                for pref_share_name_k, v in json_obj[dict["key_name"]].items():
                    pattern_pref_share_name_k = pattern.search(
                        pref_share_name_k
                    ).group()  # Extract the pattern match

                    if pattern_pref_share_name == pattern_pref_share_name_k:
                        data_extract_value = v[dict["data_extract"]]
                        if field == "preferred_shares":
                            # Convert preferred shares to int if necessary
                            data_extract_value = int(
                                data_extract_value.replace(",", "")
                            )

                        output_dict[pref_share_name].append(
                            data_extract_value
                        )  # Append data to output_dict
                        matched += 1
                        break

                if matched == 0:
                    output_dict[pref_share_name].append(
                        None
                    )  # Append None if no match found

    # Create DataFrame from output_dict and sort by "shares_type"
    df = (
        pd.DataFrame.from_dict(output_dict, orient="index", columns=columns)
        .sort_values("shares_type")
        .reset_index(drop=True)
    )

    # Function to calculate dividend percentage
    def dividend_pct_calc(pct, per_share, issue_price):
        if pct == 0:
            return per_share / issue_price
        else:
            return float(pct)

    # Function to calculate dividend per share
    def dividend_per_share_calc(pct, per_share, issue_price):
        if per_share == 0:
            return issue_price * float(pct)
        else:
            return float(per_share)

    # Apply functions to compute new DataFrame columns
    df["dividend_pct"] = df.apply(
        lambda row: dividend_pct_calc(
            row["dividend_pct"], row["dividend_per_share"], row["issue_price"]
        ),
        axis=1,
    )

    df["dividend_per_share"] = df.apply(
        lambda row: dividend_per_share_calc(
            row["dividend_pct"], row["dividend_per_share"], row["issue_price"]
        ),
        axis=1,
    )

    df["conversion_ratio"] = df.apply(
        lambda row: row["issue_price"] / row["conversion_price"],
        axis=1,
    )

    df["authorized"] = df["dividend_cumulative"]
    df["$_invested"] = df["issue_price"] * df["preferred_shares"]

    # Rename columns for clarity
    df = df.rename(
        columns={
            "participation_cap": "cap",
            "particiaption_rights": "participation",
            "dividend_cumulative": "cumulative",
        }
    )

    column_order = [
        "shares_type",
        "preferred_shares",
        "issue_price",
        "$_invested",
        "conversion_price",
        "conversion_ratio",
        "liq_pref",
        "liq_pref_order",
        "participation_rights",
        "cap",
        "dividend_pct",
        "dividend_per_share",
        "cumulative",
        "authorized",
    ]
    df = df[column_order]
    print(df.columns)

    return df, preferred_shares_list  # Return the DataFrame and preferred shares list


def convert_pandas_to_excel(bucket, df, dfs, output_dict, json_data):
    """Converts pandas dataframe to Excel workbook. Adds supporting text as comments in cells"""
    # Extract filename from the document path in the JSON data
    filename = json_data["document_txt_path"].split("/")[-1].split(".txt")[0]

    # Write the DataFrame to an in-memory buffer as an Excel file
    excel_buffer = BytesIO()
    df.to_excel(excel_buffer, index=False)

    # Load the Excel file from the buffer using openpyxl
    excel_buffer.seek(0)  # Reset buffer pointer to the beginning
    wb = load_workbook(excel_buffer)  # Load the Excel workbook
    ws = wb.active  # Get the active worksheet

    author = "coi-reader"  # Author name for comments

    # Loop through each row and column in the `dfs` DataFrame to add comments

    for row_idx in range(dfs.shape[0]):  # Loop over rows
        adj_col_idx = 1  # Adjustment to match comments to correct cell
        for col_idx in range(dfs.shape[1]):  # Loop over columns
            if col_idx in [3, 4]:
                adj_col_idx += 1  # adjusting to account for computed cells without comments ($_invested, conversion_ratio)
            comment_text = dfs.iloc[row_idx, col_idx]  # Get comment text
            len_comment_text = len(comment_text)  # Calculate comment length

            comment = Comment(text=comment_text, author=author)  # Create a comment
            # Adjust comment box dimensions based on text length
            comment.width = max(400, len_comment_text / 4)
            comment.height = max(200, len_comment_text)

            # Map Pandas DataFrame coordinates to Excel cell positions
            cell = ws.cell(
                row=row_idx + 2, column=col_idx + adj_col_idx
            )  # Offset by +2 due to Excel's 1-based index and a header row
            cell.comment = comment  # Attach the comment to the cell

    # Format specific columns (e.g., set number format to two decimal places)
    for row_idx in range(dfs.shape[0]):  # Loop over rows
        for col_idx in [2, 4]:  # Loop over specific columns
            # Map Pandas DataFrame coordinates to Excel cell positions
            cell = ws.cell(row=row_idx + 2, column=col_idx)  # Offset by +2
            cell.number_format = "#,##0"  # Set the number format

    # Adjust column widths based on header length for better readability
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        for cell in col:
            column_letter = cell.column_letter  # Get column letter
            header_length = (
                len(cell.value) if cell.value else 0
            )  # Calculate header length
            # Set column width with padding for readability
            ws.column_dimensions[column_letter].width = max(7, header_length + 1)

    # Insert 5 rows at the top of the worksheet to add additional content
    ws.insert_rows(1, 5)

    # Add content to the newly inserted rows using `output_dict`
    count = 1
    for k, v in output_dict.items():
        ws[f"A{count}"] = k  # Add field name to column A
        ws[f"B{count}"] = v["data_extract"]  # Add extracted data to column B
        comment_text = v["supporting_text"]  # Get the supporting text comment
        len_comment_text = len(comment_text)  # Calculate comment length

        # Attach comments to the cells in column B
        ws[f"B{count}"].comment = Comment(
            comment_text,
            author,
            width=max(400, len_comment_text / 4),
            height=max(200, len_comment_text),
        )
        count += 1  # Move to the next row

    # Set the width of column A for better visibility
    ws.column_dimensions["A"].width = 20

    # Save the updated workbook to an in-memory buffer
    buffer = BytesIO()
    wb.save(buffer)  # Save the workbook to the buffer
    buffer.seek(0)  # Reset buffer pointer to the beginning

    model_id = json_data["model_id"]  # Extract model ID from JSON data

    s3 = boto3.client("s3")  # Create an S3 client
    # Define S3 paths for the main Excel file and data dump
    s3_key = f"outputs/excel/{filename}/{model_id}/{filename}.xlsx"
    s3_key_data_dump = f"outputs/excel_data_dump/{filename}.xlsx"

    # Upload the Excel file to the specified S3 paths
    s3.put_object(Bucket=bucket, Key=s3_key, Body=buffer.getvalue())
    print(f"File successfully uploaded to: s3://{bucket}/{s3_key}")
    s3.put_object(Bucket=bucket, Key=s3_key_data_dump, Body=buffer.getvalue())
    print(f"File successfully uploaded to: s3://{bucket}/{s3_key_data_dump}")


def generate_support_df(bucket, json_data, preferred_shares_list):
    model_id = json_data["model_id"]  # Extract the model ID from the JSON data

    path_key = "raw_output_path"  # Key to access the raw output path in the JSON data
    columns = ["share_name"]  # Initialize the list of columns for the final DataFrame
    output_dict = {}  # Initialize an empty dictionary to store extracted data

    # Iterate over the fields to extract, defined in a global `raw_fields` dictionary
    for field, dict in raw_fields.items():
        columns.append(field)  # Add the current field to the columns list

        output_file = dict["output_file"]  # Get the output file name for the field
        file_path = json_data[output_file][
            path_key
        ]  # Retrieve the file path from JSON data
        file_content = utils.load_s3_file(
            bucket, file_path
        )  # Load the file content from S3

        if field in [
            "shares_type",
            "dividend_pct",
            "dividend_per_share",
            "dividend_cumulative",
            "liq_pref_order",
            "liq_pref",
            "participation_cap",
            "participation_rights",
        ]:
            # For each preferred share name in the list, append the file content to the output dictionary
            for pref_share_name in preferred_shares_list:
                if pref_share_name not in output_dict:
                    output_dict[pref_share_name] = [file_content]
                else:
                    output_dict[pref_share_name].append(file_content)
        else:
            # Convert the file content to a JSON object for further processing
            json_obj = utils.convert_file_to_json_obj(model_id, file_content)

            # Iterate over the preferred shares list
            for pref_share_name in preferred_shares_list:
                pattern = re.compile(regex_pattern)
                # Extract the pattern match for the current preferred share name
                pattern_pref_share_name = pattern.search(pref_share_name).group()

                matched = 0  # Initialize a variable to track if a match is found
                # Iterate over the key-value pairs in the JSON object
                for pref_share_name_k, v in json_obj[dict["key_name"]].items():
                    # Extract the pattern match for the current key in the JSON object
                    pattern_pref_share_name_k = pattern.search(
                        pref_share_name_k
                    ).group()

                    # Check if the extracted patterns match
                    if pattern_pref_share_name == pattern_pref_share_name_k:
                        data_extract_value = v[
                            dict["data_extract"]
                        ]  # Extract the required data

                        # Append the extracted data to the output dictionary
                        if pref_share_name not in output_dict:
                            output_dict[pref_share_name] = [data_extract_value]
                        else:
                            output_dict[pref_share_name].append(data_extract_value)

                        matched += 1  # Increment matched counter and break the loop
                        break

                # If no match is found, append None to the output dictionary for this preferred share name
                if matched == 0:
                    if pref_share_name not in output_dict:
                        output_dict[pref_share_name] = [None]
                    else:
                        output_dict[pref_share_name].append(None)

    # Convert the output dictionary to a DataFrame
    df = pd.DataFrame.from_dict(output_dict, orient="index").reset_index()
    df.columns = columns  # Set the DataFrame columns

    # Sort the DataFrame by "share_name" and drop the "share_name" column
    df = (
        df.sort_values("share_name").reset_index(drop=True).drop(columns=["share_name"])
    )

    column_order = [
        "shares_type",
        "preferred_shares",
        "issue_price",
        "conversion_price",
        "liq_pref",
        "liq_pref_order",
        "participation_rights",
        "participation_cap",
        "dividend_pct",
        "dividend_per_share",
        "dividend_cumulative",
    ]

    df = df[column_order]

    return df  # Return the resulting DataFrame


def extract_other_fields_dict(bucket, json_data):
    model_id = json_data["model_id"]  # Extract the model ID from the JSON data

    extract_type_list = ["raw", "precise"]  # List of extraction types to iterate over

    path_key = "raw_output_path"  # Key to access the output path in the JSON data
    # columns = ["share_name"]
    output_dict = {}  # Initialize an empty dictionary to store the extracted data

    # Iterate over the fields to extract, defined in a global `other_fields` dictionary
    for field, dict in other_fields.items():

        output_file = dict["output_file"]  # Get the output file name for the field

        file_path = json_data[output_file][
            path_key
        ]  # Retrieve the file path from JSON data
        file_content = utils.load_s3_file(
            bucket, file_path
        )  # Load the file content from S3

        interim_dict = (
            {}
        )  # Initialize a dictionary to store intermediate results for this field
        for extract_type in extract_type_list:
            if extract_type == "raw":
                # Extract supporting text based on a field in the file content
                supporting_text_extract = extract_field_value(
                    file_content, dict["supporting_text_field"]
                )
                interim_dict["supporting_text"] = (
                    supporting_text_extract  # Store the raw extraction
                )
            else:
                # Convert the file content to a JSON object
                json_obj = utils.convert_file_to_json_obj(model_id, file_content)
                interim_dict["data_extract"] = json_obj[
                    dict["key_name"]
                ]  # Extract specific data from the JSON

        output_dict[field] = (
            interim_dict  # Store the extracted data in the output dictionary
        )

    return output_dict  # Return the dictionary containing all extracted fields


def extract_field_value(json_str, field_name):
    """Function to extract value for a given field name"""

    # Escape double quotes in field name if necessary
    field_name = re.escape(field_name)
    # Regex pattern to match the specified field's value
    pattern = f'"{field_name}":\s*"(.*?)"'
    # Search for the pattern in the string
    match = re.search(pattern, json_str)
    # Return the value if a match is found
    if match:
        return match.group(1)
    else:
        return None
