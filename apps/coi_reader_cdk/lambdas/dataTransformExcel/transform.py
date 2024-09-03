import collections
import json
import re

import pandas as pd
import helpers.tranform_utils as tranform_utils
from fields import other_fields, precise_fields, raw_fields
from openpyxl import load_workbook
from openpyxl.comments import Comment
from readers import DocumentReader


def precise_output_df(path, model_id, filename, fields):
    extract = "precise"

    columns = []
    output_dict = {}
    preferred_shares_list = None
    for field, value_dict in fields.items():
        key_field = value_dict["key_name"]
        data_extract_field = value_dict["key_name"]
        columns.append(field)

        output_filename = value_dict["output_file"]
        extract_dict = tranform_utils.load_file(
            path, model_id, filename, output_filename, extract
        )

        if field == "shares_type":
            preferred_shares_list = extract_dict[value_dict["key_name"]]

            for share_type in extract_dict[value_dict["key_name"]]:
                output_dict[share_type] = [share_type]

        else:
            for pref_share_name in preferred_shares_list:

                pattern = re.compile(
                    r"\b(?:[A-Z]+\d*(?:\([A-Z0-9]+\))?(?:-[A-Z0-9]+)*|Seed)\b"
                )
                # print(pattern.search(pref_share_name).group())
                pattern_pref_share_name = pattern.search(pref_share_name).group()

                matched = 0
                for pref_share_name_k, v in extract_dict[
                    value_dict["key_name"]
                ].items():
                    pattern_pref_share_name_k = pattern.search(
                        pref_share_name_k
                    ).group()

                    if pattern_pref_share_name == pattern_pref_share_name_k:
                        data_extract_value = v[value_dict["data_extract"]]
                        if field == "preferred_shares":
                            # convert preferred shares to int
                            data_extract_value = int(
                                data_extract_value.replace(",", "")
                            )

                        output_dict[pref_share_name].append(data_extract_value)

                        matched += 1
                        break

                if matched == 0:
                    output_dict[pref_share_name].append(None)

    df = (
        pd.DataFrame.from_dict(output_dict, orient="index", columns=columns)
        .sort_values("shares_type")
        .reset_index(drop=True)
    )

    def dividend_pct_calc(pct, per_share, issue_price):

        if pct == 0:
            return per_share / issue_price
        else:
            return float(pct)

    def dividend_per_share_calc(pct, per_share, issue_price):
        if per_share == 0:
            return issue_price * float(pct)
        else:
            return float(per_share)

    df["dividend_pct"] = df.apply(
        lambda row: dividend_pct_calc(
            row["dividend_pct"], row["dividend_per_share"], row["issue_price"]
        ),
        axis=1,
    )

    df["dividend_per_share"] = df.apply(
        lambda row: dividend_per_share_calc(
            row["dividend_pct"], row["dividend_per_share"], row["issue_price"]
        ),
        axis=1,
    )

    df["conversion_ratio"] = df.apply(
        lambda row: row["issue_price"] / row["conversion_price"],
        axis=1,
    )

    df["authorized"] = df["dividend_cumulative"]
    df["$_invested"] = df["issue_price"] * df["preferred_shares"]

    df = df.rename(
        columns={
            "participation_cap": "cap",
            "particiaption_rights": "participation",
            "dividend_cumulative": "cumulative",
        }
    )

    return df, preferred_shares_list


def convert_pandas_to_excel(df, dfs, output_dict, filename):

    # Write the DataFrame to an Excel file
    excel_path = f"forge/outputs/excel/{filename}.xlsx"
    df.to_excel(excel_path, index=False)

    # Load the Excel file with openpyxl to add comments
    wb = load_workbook(excel_path)
    ws = wb.active

    author = "Bo"

    # Add comments based on the DataFrame structure
    # Loop through each row and column
    for row_idx in range(dfs.shape[0]):  # Loop over rows
        for col_idx in range(dfs.shape[1]):  # Loop over columns

            comment_text = dfs.iloc[row_idx, col_idx]
            len_comment_text = len(comment_text)

            comment = Comment(text=comment_text, author=author)
            comment.width = max(400, len_comment_text / 4)
            comment.height = max(200, len_comment_text)

            # Calculate the Excel cell position (Pandas uses 0-based indexing, Excel uses 1-based)
            cell = ws.cell(
                row=row_idx + 2, column=col_idx + 1
            )  # +2 because Excel rows start at 1 and there's a header row
            cell.comment = comment

    # Add comments based on the DataFrame structure
    # Loop through each row and column
    for row_idx in range(dfs.shape[0]):  # Loop over rows
        for col_idx in [2, 14]:  # Loop over columns
            # print(row_idx, col_idx)
            # Calculate the Excel cell position (Pandas uses 0-based indexing, Excel uses 1-based)
            cell = ws.cell(
                row=row_idx + 2, column=col_idx
            )  # +2 because Excel rows start at 1 and there's a header row

            # Specify the format you want (e.g., two decimal places)
            number_format = "#,##0"
            cell.number_format = number_format

    # Adjust the width of each column based on the length of the header
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        for cell in col:
            column_letter = cell.column_letter
            header_length = len(cell.value) if cell.value else 0
            ws.column_dimensions[column_letter].width = max(
                7, header_length + 1
            )  # Add padding for readability

    # Insert rows above the DataFrame
    ws.insert_rows(1, 5)  # Insert 2 rows at the top of the worksheet

    # Optionally, add content to the new rows
    count = 1
    for k, v in output_dict.items():
        ws[f"A{count}"] = k
        ws[f"B{count}"] = v["data_extract"]
        comment_text = v["supporting_text"]
        len_comment_text = len(comment_text)

        # print(len_comment_text, comment_text)

        # Add comments to the cells
        ws[f"B{count}"].comment = Comment(
            comment_text,
            author,
            width=max(400, len_comment_text / 4),
            height=max(200, len_comment_text),
        )

        count += 1

    ws.column_dimensions["A"].width = 20

    # Save the workbook with the changes
    wb.save(excel_path)


def support_output_df(path, preferred_shares_list, model_id, filename, fields):
    extract_type = "raw"

    columns = ["share_name"]
    output_dict = {}

    for field, value_dict in fields.items():
        # key_field = value_dict["key_name"]
        # data_extract_field = value_dict["key_name"]
        columns.append(field)

        output_filename = value_dict["output_file"]

        if field in [
            "shares_type",
            "dividend_pct",
            "dividend_per_share",
            "dividend_cumulative",
            "liq_pref_order",
            "liq_pref",
            "participation_cap",
            "participation_rights",
        ]:

            extract_file = tranform_utils.load_file(
                path, model_id, filename, output_filename, extract_type, as_json=False
            )

            for pref_share_name in preferred_shares_list:
                if pref_share_name not in output_dict:
                    output_dict[pref_share_name] = [extract_file]
                else:
                    output_dict[pref_share_name].append(extract_file)

        else:
            extract_dict = tranform_utils.load_file(
                path, model_id, filename, output_filename, extract_type, as_json=True
            )

            for pref_share_name in preferred_shares_list:
                # pattern = re.compile(r"\b[A-Z]+(?:-[A-Z\d]+)*\b")
                pattern = re.compile(
                    r"\b(?:[A-Z]+\d*(?:\([A-Z0-9]+\))?(?:-[A-Z0-9]+)*|Seed)\b"
                )
                pattern_pref_share_name = pattern.search(pref_share_name).group()

                matched = 0
                for pref_share_name_k, v in extract_dict[
                    value_dict["key_name"]
                ].items():
                    pattern_pref_share_name_k = pattern.search(
                        pref_share_name_k
                    ).group()

                    if pattern_pref_share_name == pattern_pref_share_name_k:
                        data_extract_value = v[value_dict["data_extract"]]

                        if pref_share_name not in output_dict:
                            output_dict[pref_share_name] = [data_extract_value]
                        else:
                            output_dict[pref_share_name].append(data_extract_value)

                        matched += 1
                        break

                if matched == 0:
                    if pref_share_name not in output_dict:
                        output_dict[pref_share_name] = [None]
                    else:
                        output_dict[pref_share_name].append(None)

    df = pd.DataFrame.from_dict(output_dict, orient="index").reset_index()
    df.columns = columns
    df = (
        df.sort_values("share_name").reset_index(drop=True).drop(columns=["share_name"])
    )

    return df


def extract_other_fields(path, model_id, filename, fields):

    extract_type_list = ["raw", "precise"]

    # columns = []
    output_dict = {}
    # preferred_shares_list = None
    for field, value_dict in fields.items():

        output_filename = value_dict["output_file"]

        interim_dict = {}
        for extract_type in extract_type_list:
            if extract_type == "raw":
                as_json = False
            else:
                as_json = True

            extract = tranform_utils.load_file(
                path, model_id, filename, output_filename, extract_type, as_json=as_json
            )

            if extract_type == "raw":
                supporting_text_extract = extract_field_value(
                    extract, value_dict["supporting_text_field"]
                )
                # print(field, supporting_text_extract)
                interim_dict["supporting_text"] = supporting_text_extract
            else:
                interim_dict["data_extract"] = extract[value_dict["key_name"]]

        output_dict[field] = interim_dict

    return output_dict


# Function to extract value for a given field name
def extract_field_value(json_str, field_name):
    # Escape double quotes in field name if necessary
    field_name = re.escape(field_name)
    # Regex pattern to match the specified field's value
    pattern = f'"{field_name}":\s*"(.*?)"'
    # Search for the pattern in the string
    match = re.search(pattern, json_str)
    # Return the value if a match is found
    if match:
        return match.group(1)
    else:
        return None


if __name__ == "__main__":

    model_id = "gpt-4o"

    file_list = [
        "clarity-ai 2024-08-12 COI",
        "path-ccm 2024-07-15 COI",
        "exafunction 2024-08-09 COI",
        "hungry 2024-08-09 COI",
        "invisible-technologies 2024-05-14 COI",
        "viome 2024-08-13 COI",
        "guesty 2024-03-18 COI",
        "wrapbook 2024-08-14 COI",
        "toca-football 2024-08-06 COI",
    ]

    for filename in file_list[8:]:
        print(filename)
        filepath = "forge/outputs/data_extracts"

        df, preferred_shares_list = precise_output_df(
            filepath, model_id, filename, precise_fields
        )

        dfs = support_output_df(
            filepath, preferred_shares_list, model_id, filename, raw_fields
        )
        other_fields_dict = extract_other_fields(
            filepath, model_id, filename, other_fields
        )

        convert_pandas_to_excel(df, dfs, other_fields_dict, filename)
        print(filename, ": Done")
